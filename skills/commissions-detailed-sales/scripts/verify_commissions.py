#!/usr/bin/env python3
"""
verify_commissions.py — Cross-check engine for the `commissions-detailed-sales` skill.

Recomputes monthly sales-staff commissions directly from a "Detailed sales" workbook,
using the SAME method the skill writes into Excel formulas (encaissement of the month,
net of VAT, payment-by-payment, gift vouchers excluded, Bespoke vs PAP split).

It is HEADER-DRIVEN: columns are located by their header text, never by fixed position.
Use it to validate the auditable Excel formulas the skill builds, or as a quick check.

Usage:
    python3 verify_commissions.py FILE.xlsx --month 4 [--year 2026] \
        [--vat 0.20] [--pap 0.05] [--bespoke 0.025] [--sheet "Detailed sales"] [--json]

Notes:
 * --year defaults to the current year.
 * --vat is the VAT RATE (0.20 = 20%). Use 0 for VAT-free zones (e.g. HKD).
 * Requires: pip install openpyxl
"""
import argparse, json, re, sys, unicodedata
from collections import Counter, defaultdict
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.utils import get_column_letter as L
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# Keywords used to locate the made-to-measure / petite mesure category, flagged for review.
AMBIGUOUS_TYPES = {"petite mesure", "demi mesure", "made to measure", "mtm"}
KNOWN_TYPES = {"bespoke", "stock", "cs", "c.s.", "service", "serv", "sevice",
               "servcie", "st", "pap"} | AMBIGUOUS_TYPES
PORT_HINTS = {"shipment", "shipping", "port", "livraison", "delivery",
              "frais de port", "postage", "transport"}


def norm(s):
    """lowercase, strip accents, collapse whitespace/newlines."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().lower()


def is_num(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def find_header_and_cols(ws, maxscan=20, maxcol=80):
    """Locate the header row and map logical fields -> 0-based column index, by header text."""
    hr = raw = None
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=maxscan,
                                         max_col=maxcol, values_only=True), 1):
        H = [norm(v) for v in row]
        score = sum([
            any(v == "staff" or "vendeur" in v or v == "seller" for v in H),
            any(("bespoke" in v and "stock" in v) for v in H),
            any("first payment" in v for v in H),
        ])
        if score >= 2:
            hr, raw = r, H
            break
    if hr is None:
        raise RuntimeError("Header row not found (looked for Staff / Bespoke-Stock / First payment).")

    def find(pred):
        for i, v in enumerate(raw):
            if v and pred(v):
                return i
        return None

    cols = {
        "staff":   find(lambda v: v == "staff" or "vendeur" in v or v == "seller"),
        "date":    find(lambda v: v == "date" or v == "date de vente"),
        "type":    find(lambda v: "bespoke" in v and "stock" in v),
        "category": find(lambda v: v in ("category", "categorie")),
        "model":   find(lambda v: v in ("model", "modele")),
        "p1date":  find(lambda v: "first payment" in v and "date" in v),
        "p1amt":   find(lambda v: "first payment" in v and ("amount" in v or "montant" in v)),
        "p1mode":  find(lambda v: "first payment" in v and "mode" in v),
        "p2date":  find(lambda v: "second payment" in v and "date" in v),
        "p2amt":   find(lambda v: "second payment" in v and ("amount" in v or "montant" in v)),
        "p2mode":  find(lambda v: "second payment" in v and "mode" in v),
        "p1type":  find(lambda v: "first payment" in v and "type" in v),
        # IMPORTANT: the export FLAG column is "Export or not" — NOT the price column
        # whose header happens to contain "(VAT excl. for export)".
        "export":  find(lambda v: "export" in v and ("not" in v or v == "export or not"
                                                     or v.startswith("export"))),
        "currency": find(lambda v: v in ("currency", "devise")),
    }
    return hr, cols


def compute(path, month, year, vat_rate, rate_pap=0.05, rate_bespoke=0.025,
            sheet="Detailed sales"):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    ws = wb[sheet] if sheet in wb.sheetnames else wb["Detailed sales"]
    hr, cols = find_header_and_cols(ws)

    ca = defaultdict(lambda: {"pap": 0.0, "bespoke": 0.0})
    disp = {}                       # canonical key -> display name
    rep = Counter()
    excluded_gv = 0.0
    port_flags = []
    ambiguous = Counter()
    unknown_types = Counter()
    currencies = Counter()
    missing = [k for k in ("staff", "type", "p1date", "p1amt") if cols[k] is None]

    def g(row, k):
        i = cols[k]
        return row[i] if (i is not None and i < len(row)) else None

    for r, row in enumerate(ws.iter_rows(min_row=hr + 1, values_only=True), start=hr + 1):
        staff = g(row, "staff")
        if cols["currency"] is not None:
            cv = g(row, "currency")
            if cv:
                currencies[str(cv).strip()] += 1
        tv = norm(g(row, "type"))
        is_bespoke = "bespoke" in tv
        if tv in AMBIGUOUS_TYPES:
            ambiguous[tv] += 1
        elif tv and tv not in KNOWN_TYPES:
            unknown_types[tv] += 1
        exraw = norm(g(row, "export"))
        is_export = "export" in exraw
        model_n = norm(g(row, "model"))
        cat_n = norm(g(row, "category"))
        p1type_n = norm(g(row, "p1type"))

        for tag, (dk, ak, mk) in {"p1": ("p1date", "p1amt", "p1mode"),
                                  "p2": ("p2date", "p2amt", "p2mode")}.items():
            d = g(row, dk)
            amt = g(row, ak)
            mode_n = norm(g(row, mk))
            if isinstance(d, date) and not isinstance(d, datetime):
                d = datetime(d.year, d.month, d.day)
            if not isinstance(d, datetime):
                continue
            if d.year != year or d.month != month:
                continue
            if not is_num(amt) or amt == 0:
                continue
            rep["payments_in_month"] += 1
            is_gv = ("gift voucher" in mode_n) or (tag == "p1" and p1type_n == "gift voucher")
            if is_gv:
                excluded_gv += amt
                rep["excluded_gift_voucher"] += 1
                continue
            no_vat = ("no vat" in mode_n) or ("not vat" in mode_n)
            net = amt if (vat_rate == 0 or is_export or no_vat) else amt / (1.0 + vat_rate)
            if amt < 0:
                rep["negative_lines"] += 1
            if is_export:
                rep["export_payments"] += 1
                if any(h in model_n or h in cat_n for h in PORT_HINTS):
                    port_flags.append({"row": r, "staff": str(staff), "amount": amt,
                                       "hint": (model_n or cat_n)})
                    rep["port_export_to_validate"] += 1
            display = re.sub(r"\s+", " ", str(staff).strip()) if staff else "(no staff)"
            key = display.upper()                 # case-insensitive grouping
            disp.setdefault(key, display)
            ca[key]["bespoke" if is_bespoke else "pap"] += net
            rep["payments_retained"] += 1
    wb.close()

    rows = []
    for key in sorted(ca, key=lambda k: -(ca[k]["pap"] * rate_pap + ca[k]["bespoke"] * rate_bespoke)):
        p, b = ca[key]["pap"], ca[key]["bespoke"]
        rows.append({"staff": disp[key], "ca_pap_ht": round(p, 2),
                     "ca_bespoke_ht": round(b, 2),
                     "commission": round(p * rate_pap + b * rate_bespoke, 2)})
    return {
        "header_row": hr,
        "columns": {k: (L(v + 1) if v is not None else None) for k, v in cols.items()},
        "missing_required_columns": missing,
        "currency_column_values": dict(currencies.most_common(5)) or None,
        "rates": {"pap": rate_pap, "bespoke": rate_bespoke, "vat": vat_rate},
        "period": {"month": month, "year": year},
        "staff": rows,
        "totals": {
            "ca_pap_ht": round(sum(r["ca_pap_ht"] for r in rows), 2),
            "ca_bespoke_ht": round(sum(r["ca_bespoke_ht"] for r in rows), 2),
            "commission": round(sum(r["commission"] for r in rows), 2),
        },
        "control": dict(rep),
        "excluded_gift_voucher_amount": round(excluded_gv, 2),
        "ambiguous_types_to_confirm": dict(ambiguous),
        "unknown_types_to_confirm": dict(unknown_types.most_common(10)),
        "port_export_to_validate": port_flags[:50],
    }


# ---------------------------------------------------------------------------
# Hong Kong / Asia-Pacific profile
# No Bespoke/PAP split. Base = COMBINED in-month encaissement (no tax) per staff.
# Commission = personal (turnover x per-staff rate) + store pool (tiered by % of
# store target) for salespeople; the manager gets a store-based rate instead.
# ---------------------------------------------------------------------------
DEFAULT_HK = {
    "manager": "",
    "store_target": 0,
    "personal_rate": {},          # {STAFF: rate}
    "target": {},                 # {STAFF: target$} — informational
    "pool_tiers": [[0.60, 0.003], [0.80, 0.008], [1.00, 0.015], [None, 0.022]],
    "jl_rate_reached": 0.014,
    "jl_rate_not": 0.008,
}


def commission_hk(turnover, cfg):
    """turnover: {STAFF_UPPER: combined_no_tax_turnover}. Returns the HK commission breakdown."""
    c = {**DEFAULT_HK, **cfg}
    manager = (c["manager"] or "").upper()
    store_target = float(c["store_target"] or 0)
    personal_rate = {k.upper(): v for k, v in c.get("personal_rate", {}).items()}
    tiers = c["pool_tiers"]
    store = round(sum(turnover.values()), 2)
    pct = (store / store_target) if store_target else 0.0

    def pool_rate(p):
        for thr, rate in tiers:
            if thr is None or p < thr:
                return rate
        return tiers[-1][1]

    prate = pool_rate(pct)
    pool = round(store * prate, 3)
    reached = pct >= 1.0
    jl = round(store * (c["jl_rate_reached"] if reached else c["jl_rate_not"]), 3)

    rows = []
    for s, t in sorted(turnover.items(), key=lambda kv: -kv[1]):
        if s == manager:
            personal, extra, basis = 0.0, jl, "manager (store)"
        else:
            personal, extra, basis = round(t * personal_rate.get(s, 0.0), 3), pool, "pool (store)"
        rows.append({"staff": s, "turnover": round(t, 2), "personal": personal,
                     "pool_or_manager": extra, "basis": basis, "total": round(personal + extra, 3)})
    return {
        "profile": "hk", "store_turnover": store, "store_target": store_target,
        "store_pct": round(pct, 4), "store_target_reached": reached,
        "pool_rate": prate, "pool_commission": pool,
        "manager": manager, "manager_commission": jl,
        "staff": rows, "total_commission": round(sum(r["total"] for r in rows), 3),
    }


def main():
    ap = argparse.ArgumentParser(description="Cross-check monthly sales commissions from a Detailed sales workbook.")
    ap.add_argument("file")
    ap.add_argument("--month", type=int, required=True)
    ap.add_argument("--year", type=int, default=datetime.now().year)
    ap.add_argument("--profile", choices=["europe", "hk"], default="europe",
                    help="europe = PAP/Bespoke split; hk = personal + store-pool tiers + manager")
    ap.add_argument("--vat", type=float, default=None,
                    help="VAT rate, e.g. 0.20. Defaults: europe=0.20, hk=0 (HKD).")
    ap.add_argument("--pap", type=float, default=0.05)
    ap.add_argument("--bespoke", type=float, default=0.025)
    ap.add_argument("--hk-config", help="JSON file with HK scheme params (see hk_config.example.json)")
    ap.add_argument("--sheet", default="Detailed sales")
    ap.add_argument("--json", action="store_true")
    a = ap.parse_args()
    vat = a.vat if a.vat is not None else (0.0 if a.profile == "hk" else 0.20)

    res = compute(a.file, a.month, a.year, vat, a.pap, a.bespoke, a.sheet)

    if a.profile == "hk":
        cfg = {}
        if a.hk_config:
            with open(a.hk_config, encoding="utf-8") as f:
                cfg = json.load(f)
        turnover = {r["staff"].upper(): round(r["ca_pap_ht"] + r["ca_bespoke_ht"], 2)
                    for r in res["staff"]}
        hk = commission_hk(turnover, cfg)
        if a.json:
            print(json.dumps({"base": res, "hk": hk}, ensure_ascii=False, indent=2))
            return
        print(f"HK profile — Period {a.month}/{a.year} (VAT {vat})  |  columns: {res['columns']}")
        print(f"Store turnover={hk['store_turnover']:,.2f}  target={hk['store_target']:,.0f}  "
              f"reached={hk['store_pct']:.1%} ({'YES' if hk['store_target_reached'] else 'no'})  "
              f"pool_rate={hk['pool_rate']}  manager={hk['manager'] or '(none)'}")
        print(f"\n{'STAFF':<14}{'TURNOVER':>14}{'PERSONAL':>13}{'POOL/MGR':>13}{'TOTAL':>13}  basis")
        for r in hk["staff"]:
            print(f"{r['staff'][:14]:<14}{r['turnover']:>14,.2f}{r['personal']:>13,.2f}"
                  f"{r['pool_or_manager']:>13,.2f}{r['total']:>13,.2f}  {r['basis']}")
        print(f"{'TOTAL':<14}{hk['store_turnover']:>14,.2f}{'':>13}{'':>13}{hk['total_commission']:>13,.2f}")
        print(f"\nControl: {res['control']}  | gift vouchers excluded (net): {res['excluded_gift_voucher_amount']:.2f}")
        return

    if a.json:
        print(json.dumps(res, ensure_ascii=False, indent=2))
        return
    print(f"Header row: {res['header_row']}  |  columns: {res['columns']}")
    if res["missing_required_columns"]:
        print(f"⚠ MISSING required columns: {res['missing_required_columns']} — confirm with the user.")
    cur = res["currency_column_values"]
    print(f"Currency column: {cur if cur else '(none — confirm currency with the user)'}")
    print(f"Period: {a.month}/{a.year}  | rates PAP={a.pap} Bespoke={a.bespoke} VAT={vat}")
    print(f"\n{'STAFF':<22}{'CA_PAP_HT':>15}{'CA_BESP_HT':>15}{'COMMISSION':>15}")
    for r in res["staff"]:
        print(f"{r['staff'][:22]:<22}{r['ca_pap_ht']:>15,.2f}{r['ca_bespoke_ht']:>15,.2f}{r['commission']:>15,.2f}")
    t = res["totals"]
    print(f"{'TOTAL':<22}{t['ca_pap_ht']:>15,.2f}{t['ca_bespoke_ht']:>15,.2f}{t['commission']:>15,.2f}")
    print(f"\nControl: {res['control']}")
    print(f"Gift vouchers excluded (net amount): {res['excluded_gift_voucher_amount']:.2f}")
    if res["ambiguous_types_to_confirm"]:
        print(f"⚠ Ambiguous sale types (Bespoke rate? confirm): {res['ambiguous_types_to_confirm']}")
    if res["unknown_types_to_confirm"]:
        print(f"⚠ Unrecognised sale types: {res['unknown_types_to_confirm']}")
    if res["port_export_to_validate"]:
        print(f"⚠ Port/shipping on EXPORT lines to validate manually: {len(res['port_export_to_validate'])} (see --json)")


if __name__ == "__main__":
    main()
